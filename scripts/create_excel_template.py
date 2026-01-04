"""
创建Excel测试用例模板
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 创建Excel模板
wb = Workbook()
ws = wb.active
ws.title = "测试用例"

# 设置表头
headers = ["用例ID", "用例名称", "YAML文件路径", "是否执行", "备注"]
ws.append(headers)

# 设置表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 添加示例数据
example_data = [
    ["TC001", "渠道验证接口测试", "data/test_case_01.yaml", "Y", "测试渠道验证接口"],
]

for row_data in example_data:
    ws.append(row_data)

# 调整列宽
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 30

# 保存文件
excel_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "test_cases.xlsx")
os.makedirs(os.path.dirname(excel_path), exist_ok=True)
wb.save(excel_path)
print(f"Excel模板已创建: {excel_path}")



