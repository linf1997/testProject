"""
Excel文件处理工具类
用于读取测试用例Excel文件
"""
import os
import logging
from openpyxl import load_workbook
from typing import List, Dict


class ExcelHandler:
    """Excel文件处理类"""
    
    def __init__(self, excel_path: str):
        """
        初始化Excel处理器
        :param excel_path: Excel文件路径
        """
        self.excel_path = excel_path
        self.logger = logging.getLogger(__name__)
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel文件不存在: {excel_path}")
    
    def read_test_cases(self, sheet_name: str = None) -> List[Dict]:
        """
        读取测试用例
        :param sheet_name: 工作表名称，如果为None则使用第一个工作表
        :return: 测试用例列表
        """
        try:
            workbook = load_workbook(self.excel_path)
            if sheet_name is None:
                sheet = workbook.active  # 使用活动工作表
            else:
                sheet = workbook[sheet_name]
            
            # 读取表头（第一行）
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value if cell.value else "")
            
            # 读取数据（从第二行开始）
            test_cases = []
            for row in sheet.iter_rows(min_row=2, values_only=False):
                case = {}
                for idx, cell in enumerate(row):
                    header = headers[idx] if idx < len(headers) else f"Column{idx+1}"
                    case[header] = cell.value if cell.value else ""
                
                # 跳过空行
                if any(case.values()):
                    test_cases.append(case)
            
            workbook.close()
            self.logger.info(f"成功读取 {len(test_cases)} 条测试用例")
            return test_cases
        
        except Exception as e:
            self.logger.error(f"读取Excel文件失败: {str(e)}")
            raise
    
    def get_case_by_id(self, case_id: str, sheet_name: str = None) -> Dict:
        """
        根据用例ID获取测试用例
        :param case_id: 用例ID
        :param sheet_name: 工作表名称
        :return: 测试用例字典
        """
        test_cases = self.read_test_cases(sheet_name)
        for case in test_cases:
            case_id_value = case.get("用例ID") or case.get("case_id") or case.get("用例编号")
            if str(case_id_value) == str(case_id):
                return case
        return {}


